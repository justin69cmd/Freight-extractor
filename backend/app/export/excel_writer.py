"""L6 — Mankind Excel generation (template-driven, openpyxl).

`build_workbook` is pure (rows in -> Workbook out) so it is unit-testable without
a DB. `generate_and_store` adds the review gate + persistence. Every rate sheet
carries provenance columns (Enhancement #6); a Summary sheet carries the agreement
metadata and clause digest (Enhancement #1).

openpyxl is lazy-imported so the package loads without it; only export needs it.
"""
from __future__ import annotations

from datetime import datetime, timezone
from typing import Any

from app.core.exceptions import ExportError
from app.export.templates import get_template


# --- value formatting ------------------------------------------------------ #
def _format(obj: Any, field) -> Any:
    """Resolve a template field spec against a rate-like object."""
    if isinstance(field, tuple):
        kind, attr = field
        val = getattr(obj, attr, None)
        if val is None:
            return None
        if kind == "$num":
            try:
                return float(val)
            except (TypeError, ValueError):
                return None
        if kind == "$pct":
            return round(float(val), 3)
        if kind == "$enum":
            return getattr(val, "value", str(val))
        if kind == "$cell":
            if isinstance(val, dict):
                return f"r{val.get('row')}c{val.get('col')}"
            return str(val)
        return _excel_safe(val)
    return _excel_safe(getattr(obj, field, None))


def _excel_safe(val):
    """openpyxl only accepts primitives — stringify UUIDs and other objects."""
    if val is None or isinstance(val, (str, int, float, bool)):
        return val
    from datetime import date, datetime

    if isinstance(val, (date, datetime)):
        return val
    return str(val)


def build_workbook(
    *,
    rows: list,
    metadata: Any | None,
    clauses: list,
    template_name: str,
    vendor_name: str | None = None,
    include_flagged: bool = False,
):
    """Build the Mankind workbook from canonical rate rows. Pure / DB-free."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # pragma: no cover
        raise ExportError("openpyxl not installed") from exc

    template = get_template(template_name)
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet; we add our own

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    flagged_fill = PatternFill("solid", fgColor="FCE4D6")  # tint LOW/AI rows

    # --- Summary sheet (metadata + clauses) -------------------------------- #
    _write_summary(wb, metadata, clauses, vendor_name, template_name, len(rows))

    # --- one sheet per transport mode -------------------------------------- #
    prov_cols = template["provenance_columns"] if template.get("include_provenance") else []
    by_mode: dict[str, list] = {}
    for r in rows:
        mode = getattr(r.transport_mode, "value", str(r.transport_mode))
        by_mode.setdefault(mode, []).append(r)

    for mode, spec in template["modes"].items():
        mode_rows = by_mode.get(mode, [])
        ws = wb.create_sheet(spec["sheet"])
        columns = list(spec["columns"]) + list(prov_cols)

        for c, (header, _field) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        out_row = 2
        for r in mode_rows:
            band = getattr(getattr(r, "confidence_band", None), "value", "")
            ai = bool(getattr(r, "ai_touched", False))
            is_flagged = band == "LOW" or ai or getattr(r, "rate_value", None) is None
            if is_flagged and not include_flagged:
                continue  # excluded from clean output; still in DB + review queue
            for c, (_header, field) in enumerate(columns, start=1):
                cell = ws.cell(row=out_row, column=c, value=_format(r, field))
                if is_flagged:
                    cell.fill = flagged_fill
            out_row += 1

        ws.freeze_panes = "A2"
        _autosize(ws, columns)

    return wb


def _write_summary(wb, metadata, clauses, vendor_name, template_name, rate_count) -> None:
    from openpyxl.styles import Font

    ws = wb.create_sheet("Summary")
    bold = Font(bold=True)
    rows = [
        ("Mankind Freight Output", ""),
        ("Generated", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("Template", template_name),
        ("Vendor", vendor_name or getattr(metadata, "vendor_name", None) or "-"),
        ("Effective From", str(getattr(metadata, "effective_date", "") or "-")),
        ("Expiry", str(getattr(metadata, "expiry_date", "") or "-")),
        ("Payment Terms", getattr(metadata, "payment_terms", None) or "-"),
        ("Total Canonical Rates", rate_count),
        ("", ""),
        ("Clauses", ""),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=k).font = bold
        ws.cell(row=i, column=2, value=v)
    base = len(rows) + 1
    for j, cl in enumerate(clauses):
        ctype = getattr(getattr(cl, "clause_type", None), "value", "")
        ws.cell(row=base + j, column=1, value=ctype).font = bold
        ws.cell(row=base + j, column=2, value=(getattr(cl, "text", "") or "")[:300])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90


def _autosize(ws, columns) -> None:
    from openpyxl.utils import get_column_letter

    for c, (header, _f) in enumerate(columns, start=1):
        width = max(len(str(header)) + 2, 12)
        ws.column_dimensions[get_column_letter(c)].width = min(width, 40)
